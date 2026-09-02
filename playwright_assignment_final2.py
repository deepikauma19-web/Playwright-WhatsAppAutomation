"""
WhatsApp Web Automation Bot
============================
This script does 4 things, in order:

  1. LOGIN      -> Opens WhatsApp Web and waits for you to scan the QR code.
  2. READ       -> Loads your contact list from an Excel file.
  3. SEND       -> For each contact: opens their chat, sends a personalized
                   message, takes a screenshot, and reads their last 3 messages.
  4. REPORT     -> Saves what happened for every contact into a JSON file
                   and an Excel summary.

Read `main()` first for the big picture, then look at each function for details.
"""

from playwright.sync_api import sync_playwright
import openpyxl
import json
import random
import time
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# SETTINGS
# ---------------------------------------------------------------------------
CONTACTS_FILE = "/Users/deepika/Documents/GenAI Program/Playwright/contacts.xlsx"
SESSION_FOLDER = "wa_session"          # keeps you logged in between runs
SCREENSHOTS_FOLDER = Path("screenshots")
MIN_DELAY_SECONDS = 2                  # random pause range between actions,
MAX_DELAY_SECONDS = 5                  # so the bot doesn't act "robotic"
DEFAULT_COUNTRY_CODE = "+91"           # added automatically to numbers with no country code

TODAY = datetime.now().strftime("%Y-%m-%d")
JSON_REPORT_FILE = f"whatsapp_report_{TODAY}.json"
EXCEL_REPORT_FILE = f"whatsapp_report_{TODAY}.xlsx"


def pause_like_a_human():
    """Sleeps for a random amount of time so actions don't fire back-to-back."""
    time.sleep(random.uniform(MIN_DELAY_SECONDS, MAX_DELAY_SECONDS))


# ---------------------------------------------------------------------------
# STEP 1: LOGIN
# ---------------------------------------------------------------------------
def login(page):
    """
    Opens WhatsApp Web. First run: a QR code appears -> scan it with your
    phone. Every run after that: WhatsApp remembers you (session is cached
    in SESSION_FOLDER), so it skips straight to your chats.
    """
    page.goto("https://web.whatsapp.com/", wait_until="domcontentloaded", timeout=60000)

    qr_code = "canvas[aria-label='Scan me!']"
    chat_list = "div[role='grid']"
    page.wait_for_selector(f"{qr_code}, {chat_list}", timeout=60000)

    if page.locator(qr_code).is_visible():
        print("Please scan the QR code shown in the browser window with your phone.")
        page.wait_for_selector(chat_list, timeout=120000)

    print("Logged in to WhatsApp Web.")


# ---------------------------------------------------------------------------
# STEP 2: READ CONTACTS FROM EXCEL
# ---------------------------------------------------------------------------
def clean_phone_number(raw_value):
    """
    Turns whatever Excel/Sheets gave us for a phone number into a clean,
    WhatsApp-ready number with a country code.

      1. If the Phone column isn't formatted as text, Excel stores the
         number as a float (e.g. 9941419751.0). We convert floats to int
         first, so we don't end up with a stray ".0" corrupting the number.
      2. If the number doesn't already start with "+", we add
         DEFAULT_COUNTRY_CODE automatically.
    """
    if isinstance(raw_value, float) and raw_value.is_integer():
        raw_value = int(raw_value)

    text = str(raw_value).strip()
    digits_and_plus = "".join(ch for ch in text if ch.isdigit() or ch == "+")

    if not digits_and_plus.startswith("+"):
        digits_and_plus = DEFAULT_COUNTRY_CODE + digits_and_plus

    return digits_and_plus


def read_contacts(filepath=CONTACTS_FILE):
    """
    Reads contacts.xlsx and returns a list of dicts:
        {"name": "Ravi", "phone": "+919876543210", "message": "Hi {name}!"}
    Expects a header row with columns named exactly: Name, Phone, Message.
    """
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    header_row = [cell.value for cell in sheet[1]]
    name_column = header_row.index("Name")
    phone_column = header_row.index("Phone")
    message_column = header_row.index("Message")

    contacts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[name_column] is None or row[phone_column] is None:
            continue  # skip blank rows

        contacts.append({
            "name": str(row[name_column]).strip(),
            "phone": clean_phone_number(row[phone_column]),
            "message": str(row[message_column]).strip() if row[message_column] else "",
        })

    return contacts


# ---------------------------------------------------------------------------
# STEP 3: OPEN CHAT + SEND MESSAGE
# ---------------------------------------------------------------------------
def open_chat(page, phone_number):
    """
    Opens a chat directly via WhatsApp's "click to chat" link:
        https://web.whatsapp.com/send?phone=919876543210

    Returns:
        True  -> chat opened, message box is ready
        False -> WhatsApp says this number is invalid / not on WhatsApp
    """
    digits_only = phone_number.replace("+", "").replace(" ", "")
    page.goto(f"https://web.whatsapp.com/send?phone={digits_only}")

    message_box = "div[data-testid='conversation-compose-box-input']"
    invalid_number_text = "text=Phone number shared via url is invalid."
    continue_button = "text=Continue to Chat"

    # WhatsApp sometimes shows an intermediate confirmation screen with a
    # "Continue to Chat" button before actually opening the chat. If we
    # don't click it, the message box never appears.
    try:
        page.wait_for_selector(continue_button, timeout=5000)
        page.click(continue_button)
    except Exception:
        pass  # button never appeared -> WhatsApp went straight to the chat

    try:
        page.wait_for_selector(message_box, timeout=15000)
        return True
    except Exception:
        if page.locator(invalid_number_text).count() > 0:
            return False
        raise  # something else went wrong (e.g. slow network) -> let it bubble up


def send_message(page, message_text):
    """
    Types `message_text` into the open chat and presses Enter to send it.

    Confirms the send by checking whether the compose box went back to
    empty -- WhatsApp clears it the instant it accepts a message, which is
    more reliable than watching for a specific tick icon (icon names change
    across WhatsApp UI updates).

    Returns:
        True  -> compose box is empty, message was accepted for sending
        False -> compose box still has text in it (send likely failed)
    """
    message_box = page.locator("div[data-testid='conversation-compose-box-input']")
    message_box.wait_for(state="visible", timeout=10000)

    message_box.click()
    message_box.focus()
    page.wait_for_timeout(300)

    for line in message_text.split("\n"):
        page.keyboard.type(line, delay=30)
        page.keyboard.press("Shift+Enter")
    page.keyboard.press("Backspace")  # remove the one extra blank line we added above

    page.keyboard.press("Enter")

    page.wait_for_timeout(500)
    return message_box.inner_text().strip() == ""


def take_screenshot(page, contact_name):
    """Saves a screenshot of the most recently sent message bubble."""
    SCREENSHOTS_FOLDER.mkdir(exist_ok=True)
    safe_filename = "".join(ch if ch.isalnum() else "_" for ch in contact_name)
    timestamp = datetime.now().strftime("%H%M%S")
    save_path = SCREENSHOTS_FOLDER / f"{safe_filename}_{timestamp}.png"

    sent_message_bubbles = page.locator("div.message-out")
    try:
        sent_message_bubbles.last.screenshot(path=str(save_path))
    except Exception:
        page.screenshot(path=str(save_path))  # fallback: whole page

    return str(save_path)


# ---------------------------------------------------------------------------
# STEP 3b: EXTRACT LAST MESSAGES (with layered fallbacks)
# ---------------------------------------------------------------------------
def extract_last_messages(page, how_many=3):
    """
    Reads the last few messages visible in the currently open chat.

    Tries two strategies, in order:
      A) Look for structured message rows (the normal case). Direction
         (sent/received) is read from the row's data-id ("true_..." = sent,
         "false_..." = received) or, failing that, its CSS class.
      B) If no structured rows are found at all (WhatsApp's layout changed
         again), fall back to reading the raw visible text of the chat
         panel and returning its last few non-empty lines. This has no
         direction info, but guarantees we return *something* instead of
         an empty list.

    Returns a list like:
        [{"direction": "sent", "text": "Hi Ravi!"},
         {"direction": "received", "text": "Thanks!"}]
    """
    row_selectors_to_try = [
        "div.message-in, div.message-out",  # common WhatsApp Web layout
        "div[data-id]",                      # each message row has a unique id
    ]

    rows = None
    for selector in row_selectors_to_try:
        locator = page.locator(selector)
        count = locator.count()
        print(f"    [extract] tried selector '{selector}' -> {count} matches")
        if count > 0:
            rows = locator
            break

    messages = []
    if rows is not None:
        total = rows.count()
        start = max(0, total - how_many)
        for i in range(start, total):
            row = rows.nth(i)
            text = row.inner_text().strip()
            if not text:
                continue

            data_id = row.get_attribute("data-id") or ""
            row_class = row.get_attribute("class") or ""
            if data_id.startswith("true_"):
                direction = "sent"
            elif data_id.startswith("false_"):
                direction = "received"
            elif "message-out" in row_class:
                direction = "sent"
            else:
                direction = "received"

            messages.append({"direction": direction, "text": text})

    if messages:
        return messages[-how_many:]

    # Strategy B: fallback -- grab raw text from the chat panel itself.
    print("    [extract] no structured message rows found, falling back to raw chat text")
    chat_panel = page.locator("#main")
    if chat_panel.count() == 0:
        return []

    raw_text = chat_panel.inner_text()
    lines = [line.strip() for line in raw_text.split("\n") if line.strip()]
    last_lines = lines[-how_many:]
    return [{"direction": "unknown", "text": line} for line in last_lines]


# ---------------------------------------------------------------------------
# PER-CONTACT WORKFLOW
# ---------------------------------------------------------------------------
def process_one_contact(page, contact):
    """
    Runs the full flow for a single contact: open chat -> send message ->
    screenshot -> extract last messages. Any error here is caught so it
    doesn't stop the whole batch.

    Note: screenshot + extraction are attempted even if the "was it sent?"
    check comes back uncertain, since the message may well have gone
    through despite that check being imperfect -- we'd rather have partial
    data than none.
    """
    name = contact["name"]
    phone = contact["phone"]
    personalized_message = (
        contact["message"].replace("{name}", name) if contact["message"] else f"Hi {name}!"
    )

    result = {
        "name": name,
        "phone": phone,
        "message": personalized_message,
        "status": "failed",
        "error": None,
        "screenshot": None,
        "last_messages": [],
        "extraction_error": None,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }

    try:
        chat_found = open_chat(page, phone)
        if not chat_found:
            result["status"] = "not_found"
            result["error"] = "Phone number shared via url is invalid."
            return result

        pause_like_a_human()

        was_sent = send_message(page, personalized_message)
        result["status"] = "sent" if was_sent else "uncertain"
        if not was_sent:
            result["error"] = (
                "Could not confirm the compose box cleared after sending. "
                "The message may still have gone through -- check the screenshot."
            )

        pause_like_a_human()

        # Always attempt these, even if "sent" is uncertain -- partial data
        # (a screenshot, whatever messages we can read) beats none.
        result["screenshot"] = take_screenshot(page, name)

        try:
            result["last_messages"] = extract_last_messages(page, how_many=3)
            if not result["last_messages"]:
                result["extraction_error"] = "No messages found in the chat panel."
        except Exception as extraction_error:
            result["extraction_error"] = str(extraction_error)

    except Exception as error:
        result["status"] = "failed"
        result["error"] = str(error)

    return result


# ---------------------------------------------------------------------------
# STEP 4: SAVE REPORTS
# ---------------------------------------------------------------------------
def save_reports(all_results):
    """
    Writes two report files:
      - whatsapp_report_YYYY-MM-DD.json -- full details, including extracted messages
      - whatsapp_report_YYYY-MM-DD.xlsx -- one-row-per-contact summary
    """
    with open(JSON_REPORT_FILE, "w", encoding="utf-8") as f:
        json.dump({"run_date": TODAY, "contacts": all_results}, f, indent=2, ensure_ascii=False)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Summary"
    sheet.append([
        "Name", "Phone", "Status", "Message", "Screenshot",
        "Timestamp", "Error", "Last Messages", "Extraction Error",
    ])

    for result in all_results:
        last_messages_text = " | ".join(
            f"{m['direction']}: {m['text']}" for m in result["last_messages"]
        )
        sheet.append([
            result["name"],
            result["phone"],
            result["status"],
            result["message"],
            result["screenshot"] or "",
            result["timestamp"],
            result["error"] or "",
            last_messages_text,
            result["extraction_error"] or "",
        ])

    workbook.save(EXCEL_REPORT_FILE)
    print(f"\nReports saved:\n  {JSON_REPORT_FILE}\n  {EXCEL_REPORT_FILE}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    with sync_playwright() as playwright:
        context = playwright.chromium.launch_persistent_context(
            SESSION_FOLDER,
            headless=False,
            viewport={"width": 1280, "height": 800},
        )
        page = context.pages[0] if context.pages else context.new_page()

        login(page)
        contacts = read_contacts()

        all_results = []
        for contact in contacts:
            print(f"\nProcessing {contact['name']} ({contact['phone']})...")
            result = process_one_contact(page, contact)
            print(f"  -> {result['status']}")
            if result["extraction_error"]:
                print(f"  -> extraction: {result['extraction_error']}")

            all_results.append(result)
            pause_like_a_human()

        context.close()

    save_reports(all_results)


if __name__ == "__main__":
    main()